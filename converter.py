"""
OBZ Shipment Formatter - core conversion logic.

Reads a packing slip (Excel .xlsx/.xlsm/.xls, CSV, or PDF) and produces an
"OBZ Box Contents" workbook with two sheets:

    1. "Box Contents"          -> UPC | Qty | Box
    2. "Weight and Dimensions" -> Box Number | Weight | Length | Width | Height

Carton blocks on the slip look like:

    Carton <n>  …  Weight: <w>  …  PIN: <L>x<W>x<H>

followed by item rows with UPC/GTIN and Qty.

Length / Width / Height come from each carton's PIN field when filled, otherwise
from the matching DIMS workbook (Carton / Length / Width / Height). Many FBA
packing slips leave PIN blank — for those, upload the companion DIMS file.
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from io import BytesIO, StringIO
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xls", ".csv", ".pdf"}

# Column indices (1-based) inside the packing slip ----------------------------
COL_LABEL = 2     # B - "Carton" / "PO"
COL_CARTON_NO = 6  # F - carton number on the header row
COL_LICENSE = 10  # J - license plate / tracking on the header row
COL_WEIGHT = 20   # T - carton weight on the header row
COL_UPC = 23      # W - UPC/GTIN on the item rows
COL_QTY = 49      # AW - Qty on the item rows
COL_PIN = 46      # AT - PIN dimensions (e.g. "24x20x16") on the header row

DIM_PATTERN = re.compile(
    r"(\d+(?:\.\d+)?)\s*[xX\u00d7*]\s*(\d+(?:\.\d+)?)\s*[xX\u00d7*]\s*(\d+(?:\.\d+)?)"
)


@dataclass
class LineItem:
    upc: str
    qty: int


@dataclass
class Carton:
    number: int
    weight: float | int | None = None
    length: float | int | None = None
    width: float | int | None = None
    height: float | int | None = None
    license_plate: str | None = None
    items: list[LineItem] = field(default_factory=list)

    @property
    def upcs(self) -> list[str]:
        return [item.upc for item in self.items]

    @property
    def total_qty(self) -> int:
        return sum(item.qty for item in self.items)


@dataclass
class DimsRecord:
    """One carton row from a DIMS workbook."""

    carton: int | None = None
    license_plate: str | None = None
    length: float | int | None = None
    width: float | int | None = None
    height: float | int | None = None
    weight: float | int | None = None
    units: int | None = None


def _clean_digits(value) -> str:
    if value is None:
        return ""
    return "".join(ch for ch in str(value).strip() if ch.isdigit())


def gtin_to_upc(value) -> str:
    """Convert a GTIN-14 to a 12-digit UPC (trailing 12 digits)."""
    digits = _clean_digits(value)
    if len(digits) > 12:
        return digits[-12:]
    return digits


def _as_number(value):
    """Coerce a value to int or float (never a string)."""
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "")
    try:
        num = float(text)
    except ValueError:
        return None
    return int(num) if num.is_integer() else num


def parse_dimensions(value):
    """Parse a PIN dimension string like '24x20x16' into (L, W, H)."""
    if value is None:
        return (None, None, None)
    match = DIM_PATTERN.search(str(value))
    if not match:
        return (None, None, None)
    nums = [_as_number(match.group(i)) for i in (1, 2, 3)]
    return (nums[0], nums[1], nums[2])


@dataclass
class CellGrid:
    """Row/column grid with 1-based indexing, like an Excel worksheet."""

    _rows: list[list]

    @property
    def max_row(self) -> int:
        return len(self._rows)

    @property
    def max_column(self) -> int:
        return max((len(row) for row in self._rows), default=0)

    def cell_value(self, row: int, column: int):
        if row < 1 or column < 1:
            return None
        row_values = self._rows[row - 1]
        if column > len(row_values):
            return None
        value = row_values[column - 1]
        if value == "":
            return None
        return value


def _normalize_extension(filename: str | None) -> str:
    if not filename:
        return ".xlsx"
    return Path(filename).suffix.lower()


def _read_bytes(source) -> bytes:
    if isinstance(source, (bytes, bytearray)):
        return bytes(source)
    if hasattr(source, "read"):
        data = source.read()
        if hasattr(source, "seek"):
            source.seek(0)
        return data
    return Path(source).read_bytes()


def _grid_from_openpyxl(ws) -> CellGrid:
    rows: list[list] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return CellGrid(rows)


def _grid_from_xls(data: bytes) -> CellGrid:
    import xlrd

    book = xlrd.open_workbook(file_contents=data)
    sheet = book.sheet_by_index(0)
    rows: list[list] = []
    for r in range(sheet.nrows):
        row_values = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_EMPTY:
                row_values.append(None)
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                value = cell.value
                row_values.append(int(value) if value == int(value) else value)
            else:
                row_values.append(cell.value)
        rows.append(row_values)
    return CellGrid(rows)


def _grid_from_csv(data: bytes) -> CellGrid:
    text = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = data.decode("utf-8", errors="replace")

    rows: list[list] = []
    for row in csv.reader(StringIO(text)):
        rows.append(row)
    return CellGrid(rows)


def _extract_pdf_text(data: bytes) -> str:
    """Extract plain text from a packing-slip PDF."""
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover
        raise ImportError(
            "PDF support requires pypdf. Install it with: pip install pypdf"
        ) from exc

    reader = PdfReader(BytesIO(data))
    parts: list[str] = []
    for page in reader.pages:
        parts.append(page.extract_text() or "")
    return "\n".join(parts)


def parse_packing_slip_pdf(source) -> list[Carton]:
    """Parse a packing-slip PDF into cartons using Carton / PIN / Qty text markers."""
    data = _read_bytes(source)
    text = _extract_pdf_text(data)
    if not text.strip():
        raise ValueError("Could not extract text from that PDF.")

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    cartons: list[Carton] = []
    current: Carton | None = None

    carton_re = re.compile(r"\bCarton\b\s+(\d+)\b", re.IGNORECASE)
    weight_re = re.compile(r"\bWeight\s*:?\s*([\d.]+)", re.IGNORECASE)
    pin_re = re.compile(r"\bPIN\s*:?\s*" + DIM_PATTERN.pattern, re.IGNORECASE)
    # 12–14 digit codes (UPC/GTIN), optional qty nearby
    upc_re = re.compile(r"\b(\d{12,14})\b")
    qty_near_re = re.compile(r"(?:Qty|Quantity)\s*:?\s*(\d+)", re.IGNORECASE)

    for i, line in enumerate(lines):
        carton_match = carton_re.search(line)
        if carton_match:
            # Header lines often span nearby lines — gather a small window.
            window = " ".join(lines[i : i + 4])
            weight_match = weight_re.search(window)
            pin_match = pin_re.search(window)
            if not pin_match:
                pin_match = DIM_PATTERN.search(window)

            length = width = height = None
            if pin_match:
                if pin_match.lastindex and pin_match.lastindex >= 3:
                    length = _as_number(pin_match.group(1))
                    width = _as_number(pin_match.group(2))
                    height = _as_number(pin_match.group(3))
                else:
                    length, width, height = parse_dimensions(pin_match.group(0))

            # License plate: long digit token on the header window
            license_plate = None
            for token in re.findall(r"\b\d{8,}\b", window):
                if len(token) <= 14:  # avoid SSCC-18 / GTIN
                    license_plate = token
                    break

            current = Carton(
                number=int(carton_match.group(1)),
                weight=_as_number(weight_match.group(1)) if weight_match else None,
                length=length,
                width=width,
                height=height,
                license_plate=license_plate,
            )
            cartons.append(current)
            continue

        if current is None:
            continue

        # Skip obvious header rows
        lower = line.lower()
        if "upc" in lower and "gtin" in lower:
            continue
        if lower.startswith("po") and "pick" in lower:
            continue

        upc_matches = upc_re.findall(line)
        if not upc_matches:
            continue

        qty = 1
        qty_match = qty_near_re.search(line)
        if qty_match:
            qty = int(qty_match.group(1))
        else:
            # Trailing small integer on the line is often Qty
            trailing = re.search(r"\b(\d{1,4})\s*$", line)
            if trailing:
                candidate = int(trailing.group(1))
                # Avoid treating part of a long code / year-like values as qty
                if candidate <= 100 and trailing.group(1) not in upc_matches[-1]:
                    qty = candidate

        upc = gtin_to_upc(upc_matches[-1])
        if len(upc) >= 12:
            current.items.append(LineItem(upc=upc, qty=qty))

    return cartons


def load_cell_grid(source, filename: str | None = None) -> CellGrid:
    """Load a packing slip from Excel (.xlsx/.xlsm/.xls) or CSV into a grid."""
    ext = _normalize_extension(filename)
    if ext not in SUPPORTED_EXTENSIONS - {".pdf"}:
        supported = ", ".join(sorted(SUPPORTED_EXTENSIONS))
        raise ValueError(f"Unsupported file type '{ext}'. Use one of: {supported}")

    data = _read_bytes(source)

    if ext == ".csv":
        return _grid_from_csv(data)
    if ext == ".xls":
        return _grid_from_xls(data)

    wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
    return _grid_from_openpyxl(wb.active)


def _pin_from_row_values(row_values: list) -> tuple:
    """Find PIN dimensions on a carton header row."""
    # 1) Same cell: "PIN: 24x20x16"
    for value in row_values:
        if value is None:
            continue
        text = str(value).strip()
        if text.upper().startswith("PIN"):
            length, width, height = parse_dimensions(text)
            if length is not None:
                return length, width, height

    # 2) Value in the cell after a "PIN:" label
    for idx, value in enumerate(row_values):
        if value is None:
            continue
        if str(value).strip().upper().rstrip(":") == "PIN":
            for later in row_values[idx + 1 :]:
                length, width, height = parse_dimensions(later)
                if length is not None:
                    return length, width, height
            break

    # 3) Dedicated PIN value column (AT)
    if len(row_values) >= COL_PIN:
        length, width, height = parse_dimensions(row_values[COL_PIN - 1])
        if length is not None:
            return length, width, height

    # 4) Any LxWxH token on the row
    for value in row_values:
        length, width, height = parse_dimensions(value)
        if length is not None:
            return length, width, height

    return (None, None, None)


def _scan_carton_fields(grid: CellGrid, row: int):
    """Read carton number, license plate, weight, and PIN dimensions from a header row."""
    carton_no = _as_number(grid.cell_value(row, COL_CARTON_NO))
    license_plate = grid.cell_value(row, COL_LICENSE)
    if license_plate is not None:
        license_plate = str(license_plate).strip() or None
    weight = _as_number(grid.cell_value(row, COL_WEIGHT))

    row_values = [
        grid.cell_value(row, col)
        for col in range(1, grid.max_column + 1)
    ]
    length, width, height = _pin_from_row_values(row_values)

    if weight is None:
        for idx, value in enumerate(row_values):
            if value and "weight" in str(value).lower():
                for later in row_values[idx + 1 :]:
                    parsed = _as_number(later)
                    if parsed is not None:
                        weight = parsed
                        break
                break

    if carton_no is None:
        seen_carton = False
        for value in row_values:
            if value is None:
                continue
            text = str(value).strip().lower()
            if text == "carton":
                seen_carton = True
                continue
            if seen_carton:
                parsed = _as_number(value)
                if parsed is not None:
                    carton_no = parsed
                    break

    if license_plate is None:
        for value in row_values:
            if value is None:
                continue
            text = str(value).strip()
            digits = _clean_digits(text)
            if len(digits) >= 8 and digits == text and len(digits) <= 14:
                license_plate = text
                break

    return carton_no, license_plate, weight, length, width, height


def _detect_header_column(grid: CellGrid, row: int, *names: str) -> int | None:
    targets = [n.lower() for n in names]
    for col in range(1, grid.max_column + 1):
        value = grid.cell_value(row, col)
        if value is None:
            continue
        text = str(value).strip().lower()
        if any(t == text or t in text for t in targets):
            return col
    return None


def parse_packing_slip_grid(grid: CellGrid) -> list[Carton]:
    """Parse a cell grid into an ordered list of Cartons."""
    cartons: list[Carton] = []
    current: Carton | None = None
    col_upc = COL_UPC
    col_qty = COL_QTY

    for row in range(1, grid.max_row + 1):
        label = grid.cell_value(row, COL_LABEL)
        label_str = str(label).strip().lower() if label is not None else ""

        if label_str == "po":
            detected_upc = _detect_header_column(grid, row, "upc", "upc/gtin", "gtin")
            if detected_upc is not None:
                col_upc = detected_upc
            detected_qty = _detect_header_column(grid, row, "qty", "quantity")
            if detected_qty is not None:
                col_qty = detected_qty
            continue

        if label_str == "carton":
            carton_no, license_plate, weight, length, width, height = _scan_carton_fields(
                grid, row
            )
            current = Carton(
                number=int(carton_no) if carton_no is not None else len(cartons) + 1,
                weight=weight,
                length=length,
                width=width,
                height=height,
                license_plate=license_plate,
            )
            cartons.append(current)
            continue

        raw_upc = grid.cell_value(row, col_upc)
        digits = _clean_digits(raw_upc)
        if current is not None and len(digits) >= 12:
            qty_val = _as_number(grid.cell_value(row, col_qty))
            qty = int(qty_val) if qty_val is not None and qty_val > 0 else 1
            current.items.append(LineItem(upc=gtin_to_upc(raw_upc), qty=qty))

    return cartons


def parse_packing_slip(source, filename: str | None = None) -> list[Carton]:
    """Parse a packing slip (Excel/CSV/PDF) into an ordered list of Cartons."""
    ext = _normalize_extension(filename)
    if ext == ".pdf":
        return parse_packing_slip_pdf(source)
    grid = load_cell_grid(source, filename)
    return parse_packing_slip_grid(grid)


def _find_header_row(
    grid: CellGrid, required: set[str], scan_rows: int = 15
) -> tuple[int, dict[str, int]] | None:
    """Locate a header row containing the required column names."""
    required_norm = {name.lower() for name in required}
    for row in range(1, min(grid.max_row, scan_rows) + 1):
        mapping: dict[str, int] = {}
        for col in range(1, grid.max_column + 1):
            value = grid.cell_value(row, col)
            if value is None:
                continue
            key = str(value).strip().lower()
            if key:
                mapping[key] = col
        if required_norm.issubset(mapping.keys()):
            return row, mapping
    return None


def parse_dims_file(source, filename: str | None = None) -> list[DimsRecord]:
    """Parse a DIMS workbook (Length / Width / Height / Weight per carton)."""
    grid = load_cell_grid(source, filename)
    found = _find_header_row(grid, {"carton", "length", "width", "height"})
    if found is None:
        raise ValueError(
            "Could not find a DIMS header row with Carton / Length / Width / Height."
        )

    header_row, cols = found
    col_carton = cols["carton"]
    col_length = cols["length"]
    col_width = cols["width"]
    col_height = cols["height"]
    col_license = cols.get("license plate")
    col_weight = cols.get("weight")
    col_units = cols.get("units")

    records: list[DimsRecord] = []
    for row in range(header_row + 1, grid.max_row + 1):
        carton_no = _as_number(grid.cell_value(row, col_carton))
        length = _as_number(grid.cell_value(row, col_length))
        width = _as_number(grid.cell_value(row, col_width))
        height = _as_number(grid.cell_value(row, col_height))
        if carton_no is None:
            continue

        license_plate = None
        if col_license is not None:
            raw = grid.cell_value(row, col_license)
            if raw is not None:
                license_plate = str(raw).strip() or None

        weight = _as_number(grid.cell_value(row, col_weight)) if col_weight else None
        units_val = _as_number(grid.cell_value(row, col_units)) if col_units else None
        units = int(units_val) if units_val is not None else None

        records.append(
            DimsRecord(
                carton=int(carton_no),
                license_plate=license_plate,
                length=length,
                width=width,
                height=height,
                weight=weight,
                units=units,
            )
        )
    return records


def apply_dims_to_cartons(cartons: list[Carton], dims: list[DimsRecord]) -> int:
    """Merge DIMS L/W/H onto cartons. Match by license plate, then carton number.

    Only fills dimensions that are still missing (PIN takes priority).
    Returns the number of cartons that received at least one dimension.
    """
    by_license = {d.license_plate: d for d in dims if d.license_plate}
    by_number = {d.carton: d for d in dims if d.carton is not None}

    updated = 0
    for carton in cartons:
        match = None
        if carton.license_plate and carton.license_plate in by_license:
            match = by_license[carton.license_plate]
        elif carton.number in by_number:
            match = by_number[carton.number]
        if match is None:
            continue

        changed = False
        if carton.length is None and match.length is not None:
            carton.length = match.length
            changed = True
        if carton.width is None and match.width is not None:
            carton.width = match.width
            changed = True
        if carton.height is None and match.height is not None:
            carton.height = match.height
            changed = True
        if changed:
            updated += 1
    return updated


def build_box_contents_workbook(cartons: list[Carton]) -> openpyxl.Workbook:
    """Create the OBZ Box Contents workbook matching the reference template."""
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Box Contents"

    bold = Font(name="Calibri", size=11, bold=True)
    normal = Font(name="Calibri", size=11, bold=False)

    headers = [("A", "UPC", "0.00"), ("B", "Qty", "0"), ("C", "Box", "General")]
    for col, text, fmt in headers:
        cell = ws1[f"{col}1"]
        cell.value = text
        cell.font = bold
        cell.number_format = fmt

    row = 2
    for carton in cartons:
        for item in carton.items:
            a = ws1.cell(row=row, column=1, value=item.upc)
            a.number_format = "0.00"
            a.font = normal

            b = ws1.cell(row=row, column=2, value=int(item.qty))
            b.number_format = "0"
            b.font = normal

            c = ws1.cell(row=row, column=3, value=int(carton.number))
            c.number_format = "General"
            c.font = normal
            row += 1

    ws1.column_dimensions["A"].width = 13.11
    ws1.column_dimensions["B"].width = 8.89

    ws2 = wb.create_sheet("Weight and Dimensions")
    for idx, text in enumerate(
        ["Box Number", "Weight", "Length", "Width", "Height"], start=1
    ):
        cell = ws2.cell(row=1, column=idx, value=text)
        cell.font = bold

    for i, carton in enumerate(cartons, start=2):
        ws2.cell(row=i, column=1, value=int(carton.number)).font = normal
        ws2.cell(row=i, column=2, value=_as_number(carton.weight)).font = normal
        ws2.cell(row=i, column=3, value=_as_number(carton.length)).font = normal
        ws2.cell(row=i, column=4, value=_as_number(carton.width)).font = normal
        ws2.cell(row=i, column=5, value=_as_number(carton.height)).font = normal

    ws2.column_dimensions["A"].width = 11.44
    ws2.column_dimensions["C"].width = 8.78

    return wb


def convert(
    source,
    filename: str | None = None,
    dims_source=None,
    dims_filename: str | None = None,
) -> tuple[openpyxl.Workbook, list[Carton]]:
    cartons = parse_packing_slip(source, filename=filename)
    if dims_source is not None:
        dims = parse_dims_file(dims_source, filename=dims_filename)
        apply_dims_to_cartons(cartons, dims)
    wb = build_box_contents_workbook(cartons)
    return wb, cartons


def workbook_to_bytes(wb: openpyxl.Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
